"""Read Excel template files to understand expected output format."""
import openpyxl

for fname in ['data/result1.xlsx', 'data/result2.xlsx', 'data/result4.xlsx']:
    print(f'\n{"="*60}')
    print(f'FILE: {fname}')
    print('='*60)
    wb = openpyxl.load_workbook(fname)
    for name in wb.sheetnames:
        ws = wb[name]
        print(f'Sheet: "{name}", rows={ws.max_row}, cols={ws.max_column}')
        # Print first 10 rows
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=min(10, ws.max_row), values_only=True)):
            print(f'  Row {i+1}: {list(row)}')
        # Check last few rows too
        if ws.max_row > 10:
            print('  ...')
            for i, row in enumerate(ws.iter_rows(min_row=ws.max_row-3, max_row=ws.max_row, values_only=True)):
                print(f'  Row {ws.max_row-3+i}: {list(row)}')
        # Check column types/count
        print(f'  Column dimensions: {ws.dimensions}')
        # Print first row (headers) in detail
        if ws.max_row >= 1:
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            print(f'  Headers: {headers}')
            print(f'  Number of header columns: {len([h for h in headers if h is not None])}')
