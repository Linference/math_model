"""
utils.py — 2024 CUMCM Problem A: Bench Dragon
Common utility functions for all sub-problems.

Mathematical core:
  - Archimedean spiral: r(theta) = b * theta,  b = p/(2*pi)
  - Arc length: s(theta1, theta2) = (b/2) * [theta*sqrt(theta^2+1) + arcsinh(theta)]_{theta1}^{theta2}
  - Handle chain: recursive bisection to place 224 handles on the spiral
  - Speed computation: central differencing
  - SAT collision detection for bench rectangles

Usage: import from solve_qN.py scripts.
All internal units: SI (meters, seconds).
"""

import numpy as np
from scipy.optimize import bisect

# --- Geometry Parameters ---
N_HANDLES = 224        # Total handle points (head front to tail back)
N_BENCHES = 223         # Total bench sections
HEAD_BOARD_LEN = 3.41   # m (341 cm)
BODY_BOARD_LEN = 2.20   # m (220 cm)
BOARD_WIDTH = 0.30      # m (30 cm)
HOLE_OFFSET = 0.275     # m (27.5 cm from board end)

# Effective lengths (handle-to-handle distance)
L_HEAD = HEAD_BOARD_LEN - 2 * HOLE_OFFSET  # 2.86 m (handle 1-2)
L_BODY = BODY_BOARD_LEN - 2 * HOLE_OFFSET  # 1.65 m (handle i to i+1 for i>=2)

# Build array of effective lengths for sequential handle pairs
def get_L_array():
    """Return array L[i] = distance between handle i and handle i+1, i=0..222."""
    L = np.full(N_BENCHES, L_BODY)
    L[0] = L_HEAD  # head bench
    return L


# --- Spiral Parameterization ---
def b_from_p(p):
    """b = p / (2*pi)"""
    return p / (2.0 * np.pi)


def spiral_position(theta, b):
    """Return (x, y) in meters for spiral point at angle theta (rad), param b."""
    r = b * theta
    x = r * np.cos(theta)
    y = r * np.sin(theta)
    return np.column_stack([x, y]) if np.ndim(theta) > 0 else np.array([x, y])


def spiral_derivative(theta, b):
    """
    Return (dx/dtheta, dy/dtheta) analytically.
    d/dtheta [b*theta*cos(theta)] = b*cos(theta) - b*theta*sin(theta)
    d/dtheta [b*theta*sin(theta)] = b*sin(theta) + b*theta*cos(theta)
    """
    dx = b * (np.cos(theta) - theta * np.sin(theta))
    dy = b * (np.sin(theta) + theta * np.cos(theta))
    return np.column_stack([dx, dy]) if np.ndim(theta) > 0 else np.array([dx, dy])


def spiral_arc_length(theta1, theta2, b):
    """
    Compute arc length of Archimedean spiral between theta1 and theta2.
    s = (b/2) * [F(theta2) - F(theta1)]
    where F(theta) = theta * sqrt(theta^2+1) + arcsinh(theta)
    """
    def F(t):
        return t * np.sqrt(t**2 + 1.0) + np.arcsinh(t)
    return (b / 2.0) * (F(theta2) - F(theta1))


def spiral_arc_F(theta):
    """Auxiliary function F(theta) = theta*sqrt(theta^2+1) + arcsinh(theta)."""
    return theta * np.sqrt(theta**2 + 1.0) + np.arcsinh(theta)


# --- Head Position from Arc Length ---
def find_head_theta(theta0, s_traveled, b, tol=1e-12):
    """
    Find head angle theta < theta0 such that arc length from theta to theta0 = s_traveled.
    Assumes clockwise inward motion (theta decreasing).

    Parameters:
        theta0 : initial head angle (larger)  [rad]
        s_traveled : distance traveled along spiral  [m]
        b : spiral parameter b = p/(2*pi)
        tol : convergence tolerance

    Returns theta_head such that spiral_arc_length(theta_head, theta0, b) = s_traveled.
    """
    if s_traveled <= 0:
        return theta0

    # Check if s_traveled exceeds available arc (to theta=0)
    s_max = spiral_arc_length(0.0, theta0, b)
    if s_traveled >= s_max:
        return 0.0  # reached center (theoretical limit)

    # Bisection: find theta where arc(theta, theta0) = s
    theta_low = 0.0
    theta_high = theta0

    # Use analytical bound: for small s, theta is close to theta0
    # arc ≈ b*sqrt(theta^2+1) * (theta0-theta) for small delta
    # theta_est = theta0 - s / (b * sqrt(theta0^2+1))
    theta_guess = theta0 - s_traveled / (b * np.sqrt(theta0**2 + 1.0))
    theta_guess = max(1e-6, theta_guess)

    for _ in range(200):
        theta_mid = (theta_low + theta_high) / 2.0
        s_mid = spiral_arc_length(theta_mid, theta0, b)
        if abs(s_mid - s_traveled) < tol:
            return theta_mid
        if s_mid < s_traveled:
            theta_high = theta_mid
        else:
            theta_low = theta_mid

    return (theta_low + theta_high) / 2.0


# --- Handle Chain Computation ---
def distance_between_spiral_points(theta1, theta2, b):
    """
    Euclidean distance between two points on Archimedean spiral.
    d^2 = b^2 * [theta1^2 + theta2^2 - 2*theta1*theta2*cos(theta1-theta2)]
    """
    d2 = b**2 * (theta1**2 + theta2**2 - 2.0 * theta1 * theta2 * np.cos(theta1 - theta2))
    return np.sqrt(max(0.0, d2))


def find_next_theta(theta_i, L, b, tol=1e-12):
    """
    Given theta_i (handle i on spiral), find theta_{i+1} > theta_i
    such that Euclidean distance = L.

    For clockwise spiral-in: tail is behind head, so theta increases along the chain.
    theta_{i+1} > theta_i.

    Bisection search range is [theta_i, theta_i + dtheta_max].
    """
    if L <= 0:
        return theta_i

    # Estimate angular step from linearized distance
    # d ≈ b * sqrt(theta^2 + 1) * dtheta  for small dtheta
    dtheta_est = L / (b * np.sqrt(theta_i**2 + 1.0))

    # Search range
    theta_low = theta_i
    theta_high = theta_i + max(dtheta_est * 10.0, 0.5)

    # Verify upper bound exceeds L
    d_high = distance_between_spiral_points(theta_i, theta_high, b)
    expand_count = 0
    while d_high < L and expand_count < 20:
        theta_high += dtheta_est * 5.0
        d_high = distance_between_spiral_points(theta_i, theta_high, b)
        expand_count += 1

    for _ in range(200):
        theta_mid = (theta_low + theta_high) / 2.0
        d_mid = distance_between_spiral_points(theta_i, theta_mid, b)

        if abs(d_mid - L) < tol:
            return theta_mid

        if d_mid < L:
            theta_low = theta_mid
        else:
            theta_high = theta_mid

    return (theta_low + theta_high) / 2.0


def compute_handle_chain(theta_head, b, L_arr=None):
    """
    Given head position theta_head, compute all 224 handle theta values
    by recursive bisection.

    theta_head is the smallest angle (head front is innermost),
    theta values increase toward the tail.

    Returns: theta array of length N_HANDLES.
    """
    if L_arr is None:
        L_arr = get_L_array()

    thetas = np.zeros(N_HANDLES)
    thetas[0] = theta_head

    for i in range(N_BENCHES):
        thetas[i + 1] = find_next_theta(thetas[i], L_arr[i], b)

    return thetas


def compute_all_positions(theta_head, b, L_arr=None):
    """
    Compute (x, y) for all 224 handles given head theta.
    Returns: positions array shape (N_HANDLES, 2).
    """
    thetas = compute_handle_chain(theta_head, b, L_arr)
    return spiral_position(thetas, b), thetas


def compute_speeds_from_history(positions_t_minus, positions_t_plus, dt):
    """
    Central difference speed for each handle.
    positions_t_minus, positions_t_plus: (N_HANDLES, 2) arrays
    dt: time step (t_plus - t_minus) / 2

    Returns: speeds array shape (N_HANDLES,).
    """
    diff = positions_t_plus - positions_t_minus
    speeds = np.sqrt(np.sum(diff**2, axis=1)) / (2.0 * dt)
    return speeds


# --- Bench Rectangle Construction ---
def bench_rectangle(handle_i, handle_j):
    """
    Get 4 corner points of a bench defined by two consecutive handle positions.

    IMPORTANT: Handles are at HOLE positions (27.5 cm from board ends).
    The actual board extends beyond the handles by HOLE_OFFSET on each end.
    Total board length = handle_distance + 2 * HOLE_OFFSET.

    The bench is a rectangle:
      - Center line extends from front_end to back_end
      - Width = BOARD_WIDTH perpendicular to center line

    Returns: corners shape (4, 2) in CCW order (front-left, front-right, back-right, back-left).
    """
    p1 = np.asarray(handle_i)   # front handle position
    p2 = np.asarray(handle_j)   # back handle position

    # Direction vector from front handle to back handle
    d = p2 - p1
    length = np.linalg.norm(d)
    if length < 1e-12:
        return np.tile(p1, (4, 1))

    dir_unit = d / length
    # Perpendicular (CCW 90 deg rotation)
    perp = np.array([-dir_unit[1], dir_unit[0]])

    half_w = BOARD_WIDTH / 2.0

    # Board ends extend HOLE_OFFSET beyond handle positions
    front_end = p1 - HOLE_OFFSET * dir_unit  # front board end
    back_end = p2 + HOLE_OFFSET * dir_unit    # back board end

    # Four corners
    corners = np.array([
        front_end - half_w * perp,  # front-left
        front_end + half_w * perp,  # front-right
        back_end + half_w * perp,   # back-right
        back_end - half_w * perp,   # back-left
    ])
    return corners


# --- SAT Collision Detection ---
def project_polygon(polygon, axis):
    """Project polygon onto axis, return (min, max)."""
    dots = np.dot(polygon, axis)
    return np.min(dots), np.max(dots)


def overlap_interval(a, b):
    """Check if intervals [a_min, a_max] and [b_min, b_max] overlap."""
    return not (a[1] < b[0] or b[1] < a[0])


def sat_collision_polygons(poly_a, poly_b, tol=1e-12):
    """
    Separating Axis Theorem: check if two convex polygons intersect.

    For rectangles, test 4 axes: edges of each rectangle.
    Returns True if collision detected.
    """
    # Get edges and perpendicular axes for both polygons
    edges = []

    def get_perp_axes(poly):
        n = len(poly)
        axes = []
        for i in range(n):
            edge = poly[(i + 1) % n] - poly[i]
            # Perpendicular axis
            axis = np.array([-edge[1], edge[0]])
            norm = np.linalg.norm(axis)
            if norm > 1e-15:
                axis = axis / norm
            axes.append(axis)
        return axes

    axes_a = get_perp_axes(poly_a)
    axes_b = get_perp_axes(poly_b)

    all_axes = axes_a + axes_b

    for axis in all_axes:
        proj_a = project_polygon(poly_a, axis)
        proj_b = project_polygon(poly_b, axis)
        if not overlap_interval(proj_a, proj_b):
            return False  # separating axis found

    return True  # all axes overlap → collision


def check_all_collisions(positions, verbose=False):
    """
    Check all non-adjacent bench pairs for collision using SAT.

    Adjacent benches share a handle and cannot collide.
    Returns: list of (i, j) pairs that collide, where i and j are bench indices (0-based).

    positions: (N_HANDLES, 2) array of handle positions.
    """
    collisions = []
    for i in range(N_BENCHES):
        rect_i = bench_rectangle(positions[i], positions[i + 1])
        for j in range(i + 2, N_BENCHES):  # skip adjacent
            rect_j = bench_rectangle(positions[j], positions[j + 1])
            if sat_collision_polygons(rect_i, rect_j):
                collisions.append((i, j))
                if verbose:
                    print(f"  Collision: bench {i} (handles {i}-{i+1}) vs bench {j} (handles {j}-{j+1})")
    return collisions


def has_collision(positions):
    """
    Optimized collision check: use spatial bounding-box pre-filter
    and only check bench pairs within a narrow index window.

    Strategy:
      1. Pre-compute all bench rectangles and bounding circles
      2. Only check pairs where index difference 5-30 (roughly 1-2 spiral turns)
      3. Quick center-distance filter before SAT
    """
    n = N_BENCHES
    # Pre-compute all rectangles and centers
    rects = [None] * n
    centers = np.zeros((n, 2))
    radii = np.zeros(n)  # bounding circle radius for center-distance filter

    for i in range(n):
        rects[i] = bench_rectangle(positions[i], positions[i + 1])
        centers[i] = (positions[i] + positions[i + 1]) / 2.0
        # Half-diagonal of bench rectangle as bounding radius
        board_len = np.linalg.norm(positions[i + 1] - positions[i]) + 2 * HOLE_OFFSET
        radii[i] = np.sqrt(board_len**2 + BOARD_WIDTH**2) / 2.0

    # Only check pairs within a sliding window of index differences
    for i in range(n):
        ri = rects[i]
        ci = centers[i]
        radi = radii[i]
        # Check j from i+5 to i+40 (skips adjacent + handles nearby turns)
        j_end = min(i + 40, n)
        for j in range(i + 5, j_end):
            # Quick center-distance filter
            d_centers = np.linalg.norm(ci - centers[j])
            if d_centers > radi + radii[j]:
                continue  # bounding circles don't overlap → safe

            if sat_collision_polygons(ri, rects[j]):
                return True

    return False


def collision_time_bisection(theta0, b, v_head, t_min, t_max, tol=0.01, L_arr=None):
    """
    Bisection search for the collision termination time.

    Assumes: no collision at t_min, collision at t_max.
    Finds the earliest collision time.

    Returns: (t_collision, has_collision_flag)
    If no collision even at t_max, returns (t_max, False).
    """
    if L_arr is None:
        L_arr = get_L_array()

    # First verify assumption
    s_min = v_head * t_min
    th_min = find_head_theta(theta0, s_min, b)
    pos_min, _ = compute_all_positions(th_min, b, L_arr)
    if has_collision(pos_min):
        # Try to find a lower bound without collision
        t_search = 0.0
        found_safe = False
        while t_search < t_min:
            t_search += 0.1
            s_s = v_head * t_search
            th_s = find_head_theta(theta0, s_s, b)
            pos_s, _ = compute_all_positions(th_s, b, L_arr)
            if not has_collision(pos_s):
                t_min = t_search
                found_safe = True
                break
        if not found_safe:
            # Collision from start — unusual; return a very small time
            return t_search, True

    # Check t_max
    s_max = v_head * t_max
    th_max = find_head_theta(theta0, s_max, b)
    pos_max, _ = compute_all_positions(th_max, b, L_arr)
    if not has_collision(pos_max):
        return t_max, False  # no collision found in range

    # Bisection
    for _ in range(60):
        t_mid = (t_min + t_max) / 2.0
        s_mid = v_head * t_mid
        th_mid = find_head_theta(theta0, s_mid, b)
        pos_mid, _ = compute_all_positions(th_mid, b, L_arr)

        if has_collision(pos_mid):
            t_max = t_mid
        else:
            t_min = t_mid

        if t_max - t_min < tol:
            break

    return (t_min + t_max) / 2.0, True


# --- Excel Output Helpers ---
import openpyxl

def write_result1_xlsx(output_path, times, all_positions, all_speeds):
    """
    Write result1.xlsx matching the official template format.
    - Sheet "位置": 449 rows x (1 + len(times)) cols
      Rows 2-449: alternating x, y for 224 handles
    - Sheet "速度": 225 rows x (1 + len(times)) cols
      Rows 2-225: speed for 224 handles
    """
    n_times = len(times)
    wb = openpyxl.Workbook()

    # --- Position Sheet ---
    ws_pos = wb.active
    ws_pos.title = "位置"

    # Row 1: time headers (col A blank, cols B+ = "t s")
    for j, t in enumerate(times):
        ws_pos.cell(row=1, column=j + 2, value=f"{t} s")

    # Handle labels
    handle_labels = []
    handle_labels.append(("龙头x (m)", "龙头y (m)"))
    for k in range(1, 222):  # 第1节龙身 to 第221节龙身
        handle_labels.append((f"第{k}节龙身x (m)", f"第{k}节龙身y (m)"))
    handle_labels.append(("龙尾x (m)", "龙尾y (m)"))
    handle_labels.append(("龙尾后x (m)", "龙尾后y (m)"))

    for i in range(N_HANDLES):
        row_x = 2 * i + 2
        row_y = 2 * i + 3
        ws_pos.cell(row=row_x, column=1, value=handle_labels[i][0])
        ws_pos.cell(row=row_y, column=1, value=handle_labels[i][1])
        for j in range(n_times):
            ws_pos.cell(row=row_x, column=j + 2, value=round(float(all_positions[j, i, 0]), 6))
            ws_pos.cell(row=row_y, column=j + 2, value=round(float(all_positions[j, i, 1]), 6))

    # --- Speed Sheet ---
    ws_spd = wb.create_sheet("速度")

    for j, t in enumerate(times):
        ws_spd.cell(row=1, column=j + 2, value=f"{t} s")

    spd_labels = ["龙头 (m/s)"]
    for k in range(1, 222):
        spd_labels.append(f"第{k}节龙身 (m/s)")
    spd_labels.append("龙尾 (m/s)")
    spd_labels.append("龙尾后 (m/s)")

    for i in range(N_HANDLES):
        row = i + 2
        ws_spd.cell(row=row, column=1, value=spd_labels[i])
        for j in range(n_times):
            ws_spd.cell(row=row, column=j + 2, value=round(float(all_speeds[j, i]), 6))

    wb.save(output_path)
    print(f"[OK] Written {output_path}")


def write_result2_xlsx(output_path, positions, speeds):
    """
    Write result2.xlsx: single snapshot.
    225 rows x 4 cols: handle label, x, y, speed.
    """
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.cell(row=1, column=2, value="横坐标x (m)")
    ws.cell(row=1, column=3, value="纵坐标y (m)")
    ws.cell(row=1, column=4, value="速度 (m/s)")

    spd_labels = ["龙头"]
    for k in range(1, 222):
        spd_labels.append(f"第{k}节龙身")
    spd_labels.append("龙尾")
    spd_labels.append("龙尾后")

    for i in range(N_HANDLES):
        row = i + 2
        ws.cell(row=row, column=1, value=spd_labels[i])
        ws.cell(row=row, column=2, value=round(float(positions[i, 0]), 6))
        ws.cell(row=row, column=3, value=round(float(positions[i, 1]), 6))
        ws.cell(row=row, column=4, value=round(float(speeds[i]), 6))

    wb.save(output_path)
    print(f"[OK] Written {output_path}")


def write_result4_xlsx(output_path, times, all_positions, all_speeds):
    """
    Write result4.xlsx: same structure as result1 but with different time range.
    - Sheet "位置": 449 rows x (1 + len(times)) cols
    - Sheet "速度": 225 rows x (1 + len(times)) cols
    """
    n_times = len(times)
    wb = openpyxl.Workbook()

    # --- Position Sheet ---
    ws_pos = wb.active
    ws_pos.title = "位置"

    for j, t in enumerate(times):
        ws_pos.cell(row=1, column=j + 2, value=f"{t} s")

    handle_labels = []
    handle_labels.append(("龙头x (m)", "龙头y (m)"))
    for k in range(1, 222):
        handle_labels.append((f"第{k}节龙身x (m)", f"第{k}节龙身y (m)"))
    handle_labels.append(("龙尾x (m)", "龙尾y (m)"))
    handle_labels.append(("龙尾后x (m)", "龙尾后y (m)"))

    for i in range(N_HANDLES):
        row_x = 2 * i + 2
        row_y = 2 * i + 3
        ws_pos.cell(row=row_x, column=1, value=handle_labels[i][0])
        ws_pos.cell(row=row_y, column=1, value=handle_labels[i][1])
        for j in range(n_times):
            ws_pos.cell(row=row_x, column=j + 2, value=round(float(all_positions[j, i, 0]), 6))
            ws_pos.cell(row=row_y, column=j + 2, value=round(float(all_positions[j, i, 1]), 6))

    # --- Speed Sheet ---
    ws_spd = wb.create_sheet("速度")

    for j, t in enumerate(times):
        ws_spd.cell(row=1, column=j + 2, value=f"{t} s")

    spd_labels = ["龙头 (m/s)"]
    for k in range(1, 222):
        spd_labels.append(f"第{k}节龙身 (m/s)")
    spd_labels.append("龙尾 (m/s)")
    spd_labels.append("龙尾后 (m/s)")

    for i in range(N_HANDLES):
        row = i + 2
        ws_spd.cell(row=row, column=1, value=spd_labels[i])
        for j in range(n_times):
            ws_spd.cell(row=row, column=j + 2, value=round(float(all_speeds[j, i]), 6))

    wb.save(output_path)
    print(f"[OK] Written {output_path}")


# --- Simulation Loop Helper ---
def simulate_time_range(theta0, b, v_head, times, L_arr=None):
    """
    Run full simulation for given time points.

    Parameters:
        theta0: initial head angle
        b: spiral parameter
        v_head: head speed (m/s)
        times: array of time points

    Returns:
        all_positions: (len(times), N_HANDLES, 2) array
        all_speeds: (len(times), N_HANDLES) array
        all_thetas: (len(times), N_HANDLES) array
    """
    if L_arr is None:
        L_arr = get_L_array()

    n_times = len(times)
    all_positions = np.zeros((n_times, N_HANDLES, 2))
    all_thetas = np.zeros((n_times, N_HANDLES))
    all_speeds = np.zeros((n_times, N_HANDLES))

    # Pre-compute all theta chains for every time step (required for positions)
    for ti, t in enumerate(times):
        s_traveled = v_head * t
        theta_cur = find_head_theta(theta0, s_traveled, b)
        thetas = compute_handle_chain(theta_cur, b, L_arr)
        all_thetas[ti] = thetas
        all_positions[ti] = spiral_position(thetas, b)

    # Compute speeds using central differences
    # Need spacing between time steps
    for ti in range(n_times):
        if ti == 0:
            # Forward difference
            dt = times[1] - times[0]
            if n_times > 1:
                diff = all_positions[ti + 1] - all_positions[ti]
                all_speeds[ti] = np.sqrt(np.sum(diff**2, axis=1)) / dt
        elif ti == n_times - 1:
            # Backward difference
            dt = times[-1] - times[-2]
            if n_times > 1:
                diff = all_positions[ti] - all_positions[ti - 1]
                all_speeds[ti] = np.sqrt(np.sum(diff**2, axis=1)) / dt
        else:
            dt = times[ti + 1] - times[ti - 1]  # should be 2*step
            diff = all_positions[ti + 1] - all_positions[ti - 1]
            all_speeds[ti] = np.sqrt(np.sum(diff**2, axis=1)) / dt

    return all_positions, all_speeds, all_thetas


# --- Key Handles for Display ---
KEY_HANDLE_INDICES = [0, 1, 51, 101, 151, 201, 223]  # 0-based

KEY_HANDLE_NAMES = [
    "龙头前把手",
    "第1节龙身前把手",
    "第51节龙身前把手",
    "第101节龙身前把手",
    "第151节龙身前把手",
    "第201节龙身前把手",
    "龙尾后把手",
]

KEY_TIME_STEPS_Q1 = [0, 60, 120, 180, 240, 300]
KEY_TIME_STEPS_Q4 = [-100, -50, 0, 50, 100]


def print_key_results(times_list, all_positions, all_speeds, times_array):
    """Print key handle data at specified times."""
    for t_target in times_list:
        ti = np.argmin(np.abs(np.array(times_array) - t_target))
        t_actual = times_array[ti]
        print(f"\n  --- t = {t_actual} s ---")
        print(f"  {'Handle':<24s} {'x (m)':<16s} {'y (m)':<16s} {'v (m/s)':<12s}")
        print(f"  {'-'*24} {'-'*16} {'-'*16} {'-'*12}")
        for idx, name in zip(KEY_HANDLE_INDICES, KEY_HANDLE_NAMES):
            x = all_positions[ti, idx, 0]
            y = all_positions[ti, idx, 1]
            v = all_speeds[ti, idx]
            print(f"  {name:<24s} {x:>15.6f}  {y:>15.6f}  {v:>11.6f}")
